import cv2
import openpyxl
import numpy as np
import math
import pandas as pd
import time
import matplotlib.pyplot as plt

capture_duration = 20

face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')
eye_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_eye.xml')

save_path = "The path where you want to save the excel with the tracking data."
# EXAMPLE: save_path = "C:/Users/John_Doe/Desktop/eye_tracking/test_0011.xlsx"

video_capture = cv2.VideoCapture(0)
frame_rate = video_capture.get(cv2.CAP_PROP_FPS)
frame_interval = 1 / frame_rate

if not video_capture.isOpened():
    print("Error: Unable to open video source")
    exit()

frame_count = 0

wb = openpyxl.Workbook()
sheet = wb.active
c1 = sheet.cell(row=1, column=1).value = "Frame Number"
c2 = sheet.cell(row=1, column=2).value = "Left Eye"
c3 = sheet.cell(row=1, column=3).value = "Right Eye"
c4 = sheet.cell(row=1, column=4).value = "Left Pupil Diameter"
c5 = sheet.cell(row=1, column=5).value = "Right Pupil Diameter"
c6 = sheet.cell(row=1, column=6).value = "Left Eye Speeds"
c7 = sheet.cell(row=1, column=7).value = "Right Eye Speeds"
c8 = sheet.cell(row=1, column=8).value = "Blink Count"

left_eye_x_coords = []
right_eye_x_coords = []
left_eye_y_coords = []
right_eye_y_coords = []
left_pupil_dia = []
right_pupil_dia = []
frame_list = []
eye_status_history = []

start_time = time.time()

def count_blinks(eye_status_history):
    blinks = 0
    previous_status = eye_status_history[0]

    for status in eye_status_history[1:]:
        if previous_status and not status:
            blinks += 1
        previous_status = status

    return blinks

def calculate_speed(x_coords, y_coords, frame_rate):
    speeds = []

    for i in range(1, len(x_coords)):
        x1, y1 = x_coords[i - 1], y_coords[i - 1]
        x2, y2 = x_coords[i], y_coords[i]
        distance_pixels = math.sqrt((x2 - x1) ** 2 + (y2 - y1) ** 2)
        distance_mm = (distance_pixels * 25.4)/141.0
        speed = distance_mm * frame_rate
        speeds.append(round(speed,2))

    return speeds

def plot_speed_graph(left_eye_speeds, right_eye_speeds, frames):
    plt.figure(figsize=(10, 6))

    plt.subplot(2,1,1)
    plt.plot(frames[1:], left_eye_speeds, label='Left Eye Speed', color='blue')
    plt.xlabel('Frame Number')
    plt.ylabel('Speed (pixels/frame)')
    plt.title('Left Eye Movement Speed Over Time')
    plt.legend()
    plt.grid(True)
    
    plt.subplot(2,1,2)
    plt.plot(frames[1:], right_eye_speeds, label='Right Eye Speed', color='red')
    plt.xlabel('Frame Number')
    plt.ylabel('Speed (pixels/frame)')
    plt.title('Right Eye Movement Speed Over Time')
    plt.legend()
    plt.grid(True)

    plt.show()

while (int (time.time()-start_time) <= capture_duration):
    ret, frame = video_capture.read()

    if not ret:
        break

    frame_count += 1

    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

    faces = face_cascade.detectMultiScale(gray, scaleFactor=1.1, minNeighbors=5, minSize=(30, 30))

    print(f"Frame {frame_count} coordinates:")

    detected_eyes = False

    for (x, y, w, h) in faces:
        roi_gray = gray[y:y+h, x:x+w]
        roi_color = frame[y:y+h, x:x+w]

        eyes = eye_cascade.detectMultiScale(roi_gray, scaleFactor=1.1, minNeighbors=10, minSize=(20, 20))

        if len(eyes) == 2:
            detected_eyes = True
            eyes = sorted(eyes, key=lambda b: b[0])

            left_eye = eyes[0]
            right_eye = eyes[1]

            left_eye_center_x = x + left_eye[0] + left_eye[2] // 2
            left_eye_center_y = y + left_eye[1] + left_eye[3] // 2
            right_eye_center_x = x + right_eye[0] + right_eye[2] // 2
            right_eye_center_y = y + right_eye[1] + right_eye[3] // 2

            left_pupil_diameter = None
            right_pupil_diameter = None
            left_pupil_diameter_mm = None
            right_pupil_diameter_mm = None

            for (ex, ey, ew, eh) in [left_eye, right_eye]:
                eye_gray = roi_gray[ey:ey+eh, ex:ex+ew]
                eye_color = roi_color[ey:ey+eh, ex:ex+ew]

                _, threshold = cv2.threshold(eye_gray, 35, 255, cv2.THRESH_BINARY_INV)
                contours, _ = cv2.findContours(threshold, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

                if contours:
                    largest_contour = max(contours, key=cv2.contourArea)
                    (x_center, y_center), radius = cv2.minEnclosingCircle(largest_contour)
                    diameter = 2 * radius

                    if ex == left_eye[0] and ey == left_eye[1]:
                        left_pupil_diameter = diameter
                        left_pupil_diameter_mm = (left_pupil_diameter * 25.4)/141.0

                    else:
                        right_pupil_diameter = diameter
                        right_pupil_diameter_mm = (right_pupil_diameter * 25.4)/141.0

            print(f"  Left eye coords: ({left_eye_center_x}, {left_eye_center_y}), Pupil Diameter: {round(left_pupil_diameter_mm, 2)}")
            print(f"  Right eye coords: ({right_eye_center_x}, {right_eye_center_y}), Pupil Diameter: {round(right_pupil_diameter_mm, 2)}")

            left_eye_x_coords.append(left_eye_center_x)
            right_eye_x_coords.append(right_eye_center_x)
            left_eye_y_coords.append(left_eye_center_y)
            right_eye_y_coords.append(right_eye_center_y)
            frame_list.append(frame_count)
            left_pupil_dia.append(left_pupil_diameter_mm)
            right_pupil_dia.append(right_pupil_diameter_mm)

            sleft_eye_x_coords = pd.Series(left_eye_x_coords)
            sright_eye_x_coords = pd.Series(right_eye_x_coords)
            sleft_eye_y_coords = pd.Series(right_eye_y_coords)
            sright_eye_y_coords =pd.Series(right_eye_y_coords)
            sframe_list = pd.Series(frame_list)
            sleft_pupil_dia = pd.Series(left_pupil_dia)
            sright_pupil_dia = pd.Series(right_pupil_dia)

            left_eye_x_coords = (sleft_eye_x_coords.interpolate()).tolist()
            right_eye_x_coords = (sright_eye_x_coords.interpolate()).tolist()
            left_eye_y_coords = (sleft_eye_y_coords.interpolate()).tolist()
            right_eye_y_coords = (sright_eye_y_coords.interpolate()).tolist()
            frame_list = (sframe_list.interpolate()).tolist()
            left_pupil_dia = (sleft_pupil_dia.interpolate()).tolist()
            right_pupil_dia = (sright_pupil_dia.interpolate()).tolist()

            left_eye_speed = calculate_speed(left_eye_x_coords, left_eye_y_coords, frame_rate)
            right_eye_speed = calculate_speed(right_eye_x_coords, right_eye_y_coords, frame_rate)

            cv2.rectangle(roi_color, (left_eye[0], left_eye[1]), (left_eye[0] + left_eye[2], left_eye[1] + left_eye[3]), (0, 255, 0), 2)
            cv2.rectangle(roi_color, (right_eye[0], right_eye[1]), (right_eye[0] + right_eye[2], right_eye[1] + right_eye[3]), (0, 255, 0), 2)

            cv2.line(frame, (left_eye_center_x - 10, left_eye_center_y), (left_eye_center_x + 10, left_eye_center_y), (0, 255, 0), 2)
            cv2.line(frame, (left_eye_center_x, left_eye_center_y - 10), (left_eye_center_x, left_eye_center_y + 10), (0, 255, 0), 2)
            cv2.line(frame, (right_eye_center_x - 10, right_eye_center_y), (right_eye_center_x + 10, right_eye_center_y), (0, 255, 0), 2)
            cv2.line(frame, (right_eye_center_x, right_eye_center_y - 10), (right_eye_center_x, right_eye_center_y + 10), (0, 255, 0), 2)
        else:
            print("Eyes not detected properly")            

    eye_status_history.append(detected_eyes)
    font = cv2.FONT_HERSHEY_COMPLEX
    for i in range(0,len(frame_list)-1): 
        j = i-1
        left_coords_and_dia = f"Left eye coords: ({left_eye_center_x}, {left_eye_center_y}), Pupil Diameter: {round(left_pupil_diameter_mm, 2)}" 
        right_coords_and_dia = f"Right eye coords: ({right_eye_center_x}, {right_eye_center_y}), Pupil Diameter: {round(right_pupil_diameter_mm, 2)}" 
        left_speed = f"Left Eye Speed: {left_eye_speed[j]}"
        right_speed = f"Right Eye Speed: {right_eye_speed[j]}"
        cv2.putText(frame,left_coords_and_dia,(0,20),font,0.5,(0,0,255),1)
        cv2.putText(frame,right_coords_and_dia,(0,35),font,0.5,(0,0,255),1)
        cv2.putText(frame,left_speed,(0,50),font,0.5,(0,0,255),1)
        cv2.putText(frame,right_speed,(0,65),font,0.5,(0,0,255),1)
    cv2.imshow('Eye Tracking', frame)

    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

blinks = count_blinks(eye_status_history)
print(f"Blink Count: {blinks}")

for i in range(0,len(frame_list)-1):
    c7 = sheet.cell(row=i+2, column=1).value = frame_list[i]
    c8 = sheet.cell(row=i+2, column=2).value = left_eye_x_coords[i]
    c9 = sheet.cell(row=i+2, column=3).value = right_eye_y_coords[i]
    c10 = sheet.cell(row=i+2, column=4).value = left_pupil_dia[i]
    c11 = sheet.cell(row=i+2, column=5).value = right_pupil_dia[i]
    c12 = sheet.cell(row=i+2, column = 6).value = left_eye_speed[i]
    c13 = sheet.cell(row=i+2, column = 7).value = right_eye_speed[i]
c14 = sheet.cell(row = 2, column = 8).value = blinks

plot_speed_graph(left_eye_speed, right_eye_speed, frame_list)

wb.save(save_path)

video_capture.release()
cv2.destroyAllWindows()